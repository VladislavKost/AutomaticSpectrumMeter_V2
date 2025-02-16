from datetime import datetime
import numpy as np

from tkinter import filedialog, messagebox, Button, Label, DISABLED
from tkinter import ttk

import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series

from RigolLib import RigolLib
from ZolixOmniUI.PyZolixOmniUI import ZolixOmniUI


class SpectralMeasurements(ZolixOmniUI):
    """Class for creating window to measure spectre of the sample using oscilloscope and monochromator"""

    def __init__(self):
        self.rigol_gateway = RigolLib.Scope()
        self.rigol_connected = False
        self.rigol_device = None
        self.oscilloscope_chanel = "ch1"

        # Первоначальные данные графика
        self.x_values = []
        self.y_values = []

        super().__init__()

    def _check_all_equipment_connected(self):
        if self.zolix_connected:
            self._enable_fields_and_buttons()
        if self.rigol_connected:
            self._enable_rigol_fields_and_buttons()

    def _enable_rigol_fields_and_buttons(self):
        self.channels_selection_box.config(state="normal")

    def _get_vertical_Rigol_scale(self):
        if self.oscilloscope_chanel == "ch1":
            self.vertical_scale = self.rigol_gateway.ch1.get_vertical_scale()
        elif self.oscilloscope_chanel == "ch2":
            self.vertical_scale = self.rigol_gateway.ch2.get_vertical_scale()

    def _connect_to_Rigol_oscilloscope(self):
        if self.rigol_gateway and self.rigol_device:
            self.rigol_gateway.manual_connect_device(self.rigol_device)
            self.rigol_gateway.auto()
            self.rigol_gateway.run()
            self.rigol_connected = True
            self._check_all_equipment_connected()
            self.rigol_connect_state.config(text="Подключено", background="#50FA1C")
            self.rigol_usb_chosen.config(state="disabled")
            self._get_vertical_Rigol_scale()

    def _create_ocsilloscope_UI(self, start_row, start_column):
        opts = self.opts

        #   Rigol USB Options
        self.rigol_usb_chosen = ttk.Combobox(
            values=self.rigol_gateway.get_available_usb_devices(),
        )

        # self.rigol_usb_chosen.current(0)
        self.rigol_usb_chosen.grid(row=start_row, column=start_column, **opts)
        self.rigol_usb_chosen.bind("<<ComboboxSelected>>", self._set_device_for_Rigol)

        # Rigol USB connect
        self.rigol_connect = Button(
            text="Подключить Rigol", command=self._connect_to_Rigol_oscilloscope
        )
        self.rigol_connect.grid(row=start_row, column=start_column + 1, **opts)

        # Rigol connect state
        self.rigol_connect_state = Label(text="Отключено", background="#F71E1E")
        self.rigol_connect_state.grid(row=start_row, column=start_column + 2, **opts)

        # Oscilloscope channel selection label
        self.channel_selection_label = Label(text="Канал осциллографа")
        self.channel_selection_label.grid(
            row=start_row + 1, column=start_column, **opts
        )

        # Oscilloscope channel selection
        self.channels_selection_box = ttk.Combobox(
            state=DISABLED,
            values=["ch1", "ch2"],
        )
        self.channels_selection_box.current(0)
        self.channels_selection_box.grid(
            row=start_row + 1, column=start_column + 1, **opts
        )
        self.channels_selection_box.bind(
            "<<ComboboxSelected>>", self._set_oscilloscope_chanel
        )

    def _create_interface(self):
        self.root.title("Управление монохроматором Zolix")  # give window name

        self.opts = {"padx": 10, "pady": 10, "ipadx": 10, "ipady": 10, "sticky": "nswe"}

        self.separator_opt = {
            "master": self.root,
            "orient": "horizontal",
            "style": "TSeparator",
            "takefocus": 1,
            "cursor": "plus",
        }

        self.style = ttk.Style()
        self.style.configure("TSeparator", background="black")

        self._create_zolix_connect_UI(start_row=0)
        ttk.Separator(**self.separator_opt).grid(
            column=0, row=1, sticky="ew", columnspan=3, padx=5, pady=5
        )

        self._create_show_cur_wl_and_grading_UI(start_row=2)
        ttk.Separator(**self.separator_opt).grid(
            column=0, row=5, sticky="ew", columnspan=3, padx=5, pady=5
        )

        # Поля и кнопки для установки диапазона длин волн
        self._create_spectrum_range_UI(start_row=6)

        # Поля и кнопки для подключения осциллографа
        self._create_ocsilloscope_UI(start_row=0, start_column=3)

        # start measurement button
        self.start_measurement_button = Button(
            state=DISABLED,
            text="Начать измерение",
            command=self._start_measurement,
        )
        self.start_measurement_button.grid(row=9, column=0, columnspan=3, **self.opts)

        # stop measurement button
        self.stop_measurement_button = Button(
            text="Завершить измерение",
            command=self._stop_measurement,
        )
        self.stop_measurement_button.grid(row=9, column=0, columnspan=3, **self.opts)
        self.stop_measurement_button.grid_remove()

    def _save_plot_excel(self):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Результаты измерения"
        sheet.cell(1, 1).value = "Длина волны"
        sheet.cell(1, 2).value = "Амплитуда"

        # Add data to the first column of the file
        for row in range(2, len(self.x_values) + 2):
            sheet.cell(row, 1).value = self.x_values[row - 2]

        # Add data to the first column of the file
        for row in range(2, len(self.y_values) + 2):
            sheet.cell(row, 2).value = self.y_values[row - 2]

        chart = ScatterChart()

        x_values = Reference(
            sheet, min_col=1, min_row=1, max_row=len(self.x_values) + 1
        )

        y_values = Reference(
            sheet, min_col=2, min_row=1, max_row=len(self.x_values) + 1
        )

        series1 = Series(y_values, x_values, title_from_data=True)

        chart.title = "Спектр излучения"
        chart.y_axis.tital = "Амплитуда, у.е"
        chart.x_axis.tital = "Длина волны, нм"

        values = Reference(
            worksheet=sheet,
            min_row=1,
            max_row=len(self.x_values) + 1,
            min_col=1,
            max_col=2,
        )

        chart.append(series1)

        sheet.add_chart(chart, "D2")

        # Open a save file dialog
        file_path = filedialog.asksaveasfilename(
            title="Выберите путь сохранения файла",
            initialfile=f"Spectrum_range_{self.initial_wl}_{self.final_wl}_step_{self.step}_{datetime.today().strftime('%d_%m_%Y')}.xlsx",
            filetypes=[("Файлы Excel", "*.xlsx")],
            defaultextension=".xlsx",
        )

        # Сохраняем файл Excel
        wb.save(file_path)

    def _save_plot_image(self):
        file_path = filedialog.asksaveasfilename(
            title="Выберите путь сохранения изображения",
            initialfile=f"Spectrum_range_{self.initial_wl}_{self.final_wl}_step_{self.step}_{datetime.today().strftime('%d_%m_%Y')}.png",
            filetypes=[("PNG files", "*.png"), ("all files", "*.*")],
            defaultextension=".png",
        )
        self.fig.savefig(file_path, dpi=1000)

    def _get_Rigol_oscillograph_average_V(self):
        if self.rigol_gateway:
            if self.oscilloscope_chanel == "ch1":
                return self.rigol_gateway.ch1.meas_Vavg()
            elif self.oscilloscope_chanel == "ch2":
                return self.rigol_gateway.ch2.meas_Vavg()

    def _get_Rigol_oscillograph_max_V(self):
        if self.rigol_gateway:
            if self.oscilloscope_chanel == "ch1":
                return self.rigol_gateway.ch1.meas_Vmax()
            elif self.oscilloscope_chanel == "ch2":
                return self.rigol_gateway.ch2.meas_Vmax()

    def _get_Rigol_oscillograph_min_V(self):
        if self.rigol_gateway:
            if self.oscilloscope_chanel == "ch1":
                return self.rigol_gateway.ch1.meas_Vmin()
            elif self.oscilloscope_chanel == "ch2":
                return self.rigol_gateway.ch2.meas_Vmin()

    def on_close(self):
        if messagebox.askokcancel("Выход", "Действительно хотите закрыть окно?"):

            self._disconnect_from_Zolix_monochromator()

            if self.rigol_connected:
                self.rigol_gateway.close_connection()

            self.root.destroy()

    def _create_save_plot_buttons(self, start_row, start_columnt):
        self.save_plot_excel_button = Button(
            self.root,
            text="Сохранить Excel",
            command=self._save_plot_excel,
        )
        self.save_plot_excel_button.grid(
            row=start_row, column=start_columnt, **self.opts
        )

        self.save_plot_image_button = Button(
            self.root,
            text="Сохранить png",
            command=self._save_plot_image,
        )
        self.save_plot_image_button.grid(
            row=start_row, column=start_columnt + 1, **self.opts
        )

    def _plot(self):
        # Очищаем прошлые данные
        self.x_values = []
        self.y_values = []

        # Блокируем кнопки
        self._disable_fields_and_buttons()
        self.start_measurement_button.grid_remove()
        self.stop_measurement_button.grid()

        if self.rigol_connected:
            # Включаем интерактивный режим
            plt.ion()

            # Создаем объект графика
            fig = plt.Figure()
            ax = fig.add_subplot(111)

            # Устанавливаем границы графика
            ax.set_xlim(self.initial_wl, self.final_wl)
            max_y_value = self._get_Rigol_oscillograph_max_V()
            ax.set_ylim(
                self._get_Rigol_oscillograph_min_V(),
                self._get_Rigol_oscillograph_max_V(),
            )

            ax.set_xlabel("Длина волны, нм")
            ax.set_ylabel("Амплитуда, у.е.")

            ax.grid(True)

            # Формируем первичную линию графика
            (line1,) = ax.plot(self.x_values, self.y_values, "b-")

            # Добавляем наш график в окно
            canvas = FigureCanvasTkAgg(fig, master=self.root)

            # Располагаем график в Tkinter окне
            canvas.get_tk_widget().grid(row=10, column=0, columnspan=3, **self.opts)

        # формируем список точек для измерения
        x_range = np.arange(self.initial_wl, self.final_wl + self.step, self.step)

        # Пробегаемся по точкам измерения и получаем данные с приборов, и обновляем график
        for x in x_range:
            if self.measuring:
                # if change_monochromator_wavelength(x):
                if self._change_monochromator_wavelength(x):
                    # Если мы не подключены к осциллографу, то пропускаем этот шаг и просто продолжаем менять длину волны на монохроматоре
                    if not self.rigol_connected:
                        continue

                    self.x_values.append(float(x))
                    new_y_value = self._get_Rigol_oscillograph_average_V()
                    self.y_values.append(new_y_value)

                    if new_y_value > max_y_value:
                        ax.set_ylim(
                            self._get_Rigol_oscillograph_min_V(),
                            new_y_value + new_y_value * 0.1,
                        )
                        max_y_value = new_y_value

                    line1.set_xdata(self.x_values)
                    line1.set_ydata(self.y_values)

                    # Обновляем график
                    fig.canvas.draw()
                    fig.canvas.flush_events()
            else:
                break

        self._enable_fields_and_buttons()
        self.stop_measurement_button.grid_remove()
        self.start_measurement_button.grid()

        if self.rigol_connected:
            # save measurement button
            self._create_save_plot_buttons(start_row=5, start_columnt=3)

            # Обновляем график
            fig.canvas.draw()
            fig.canvas.flush_events()
            self.fig = fig

    def _set_device_for_Rigol(self, event):
        val = self.rigol_usb_chosen.get()
        self.rigol_device = val

    def _set_oscilloscope_chanel(self, event):
        val = self.channels_selection_box.get()
        self.oscilloscope_chanel = val
